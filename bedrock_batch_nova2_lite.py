import argparse
import csv
import datetime as dt
import json
import os
from collections import OrderedDict
from pathlib import Path

import boto3
import openpyxl


def list_image_keys(bucket, prefix, region, limit):
    s3 = boto3.client("s3", region_name=region)
    paginator = s3.get_paginator("list_objects_v2")
    keys = []
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        for obj in page.get("Contents", []):
            key = obj["Key"]
            if key.endswith("/"):
                continue
            lower = key.lower()
            if not lower.endswith((".jpg", ".jpeg", ".png")):
                continue
            keys.append(key)
            if limit and len(keys) >= limit:
                return keys
    return keys


def load_image_keys_from_json(paths, input_prefix, limit):
    keys = []
    seen = set()
    for path in paths:
        with open(path, "r", encoding="utf-8") as f:
            items = json.load(f)
        for item in items:
            image_name = item.get("image_name")
            if not image_name:
                continue
            if image_name.startswith("s3://"):
                key = image_name.split("/", 3)[-1]
            elif "/" in image_name:
                key = image_name
            else:
                key = f"{input_prefix}{image_name}"
            if key in seen:
                continue
            seen.add(key)
            keys.append(key)
            if limit and len(keys) >= limit:
                return keys
    return keys


def list_output_jsonl_keys(bucket, prefix, region):
    s3 = boto3.client("s3", region_name=region)
    paginator = s3.get_paginator("list_objects_v2")
    keys = []
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        for obj in page.get("Contents", []):
            key = obj["Key"]
            if key.endswith(".jsonl.out"):
                keys.append(key)
    return keys


def extract_image_name_from_record(record):
    model_input = record.get("modelInput") or record.get("inputRecord", {}).get("modelInput")
    if model_input:
        for message in model_input.get("messages", []):
            for content in message.get("content", []):
                image = content.get("image")
                if not image:
                    continue
                uri = (
                    image.get("source", {})
                    .get("s3Location", {})
                    .get("uri")
                )
                if uri and "/" in uri:
                    return uri.rsplit("/", 1)[-1]
    return None


def load_processed_image_names(bucket, prefixes, region):
    s3 = boto3.client("s3", region_name=region)
    processed = set()
    for prefix in prefixes:
        for key in list_output_jsonl_keys(bucket, prefix, region):
            obj = s3.get_object(Bucket=bucket, Key=key)
            for line in obj["Body"].iter_lines():
                if not line:
                    continue
                try:
                    record = json.loads(line)
                except json.JSONDecodeError:
                    continue
                if record.get("error") or record.get("outputRecord", {}).get("error"):
                    continue
                image_name = extract_image_name_from_record(record)
                if image_name:
                    processed.add(image_name)
    return processed


def detect_format(key):
    lower = key.lower()
    if lower.endswith(".png"):
        return "png"
    return "jpeg"


def load_processed_from_csv(paths):
    processed = set()
    for path in paths:
        with open(path, "r", encoding="utf-8", newline="") as f:
            header = f.readline()
            if not header:
                continue
            columns = [c.strip() for c in header.strip().split(",")]
            try:
                idx = columns.index("image_name")
            except ValueError:
                idx = 0
            for line in f:
                parts = [p.strip() for p in line.strip().split(",")]
                if len(parts) <= idx:
                    continue
                name = parts[idx]
                if name:
                    processed.add(name)
    return processed


SYSTEM_PROMPT = (
    "You are a YouTube thumbnail analyzer. "
    "Your task is to detect whether a child is present in the thumbnail "
    "and whether the thumbnail is designed to attract children. "
    "Always respond with valid JSON only. No markdown, no explanation."
)


def build_prompt(image_name):
    return (
        f'Analyze this YouTube thumbnail (filename: "{image_name}") and return ONLY valid JSON:\n'
        "{\n"
        f'    "image_name": "{image_name}",\n'
        '    "contains_child": true or false,\n'
        '    "age_group": "infant" | "toddler" | "kids" | "teen" | "none" | "unknown",\n'
        '    "is_child_targeted": true or false,\n'
        '    "confidence": 0.0 to 1.0\n'
        "}"
    )


def build_record(record_id, s3_uri, image_name, image_format):
    return {
        "recordId": record_id,
        "modelInput": {
            "schemaVersion": "messages-v1",
            "system": [{"text": SYSTEM_PROMPT}],
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {
                            "image": {
                                "format": image_format,
                                "source": {"s3Location": {"uri": s3_uri}},
                            }
                        },
                        {"text": build_prompt(image_name)},
                    ],
                }
            ],
            "inferenceConfig": {"temperature": 0, "maxTokens": 200, "topP": 0.9},
        },
    }


def normalize_prefix(prefix):
    if prefix.endswith("/"):
        return prefix
    return f"{prefix}/"


def ensure_inference_profile(model_id):
    if not model_id:
        raise SystemExit("model_id is required")
    if not model_id.startswith("apac."):
        raise SystemExit(
            f"model_id must start with 'apac.' (e.g. apac.amazon.nova-lite-v1:0). Got: {model_id}"
        )


def video_id_from_image(image_name):
    if not image_name:
        return ""
    if "." in image_name:
        return image_name.rsplit(".", 1)[0]
    return image_name


def parse_text_payload(text_payload):
    if not text_payload:
        return None
    text_payload = text_payload.strip()
    if text_payload.startswith("```"):
        parts = text_payload.split("\n", 1)
        if len(parts) > 1:
            text_payload = parts[1]
        if text_payload.endswith("```"):
            text_payload = text_payload.rsplit("```", 1)[0]
    try:
        return json.loads(text_payload)
    except json.JSONDecodeError:
        return None


def build_book_mapping(paths):
    mapping = {}
    for path in paths:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if "video_id" not in headers or "channel_id" not in headers or "channel_display_name" not in headers:
            continue
        idx_video = headers.index("video_id")
        idx_channel_id = headers.index("channel_id")
        idx_channel_name = headers.index("channel_display_name")
        for row in ws.iter_rows(min_row=2, values_only=True):
            video_id = row[idx_video]
            if not video_id:
                continue
            if video_id in mapping:
                continue
            channel_id = row[idx_channel_id]
            channel_name = row[idx_channel_name]
            mapping[str(video_id)] = (channel_id or "", channel_name or "")
    return mapping


def build_video_info_mapping(paths):
    mapping = {}
    for path in paths:
        p = Path(path)
        if not p.exists():
            continue
        with p.open("r", encoding="utf-8") as f:
            try:
                items = json.load(f)
            except json.JSONDecodeError:
                items = []
        for item in items:
            image_name = item.get("image_name")
            video_id = item.get("video_id") or video_id_from_image(image_name or "")
            if not video_id:
                continue
            channel_id = item.get("channel_id") or item.get("channelId") or ""
            channel_name = item.get("channel_name") or item.get("channelTitle") or ""
            if video_id in mapping:
                continue
            mapping[str(video_id)] = (str(channel_id), str(channel_name))
    return mapping


def rebuild_success_csv(
    bucket,
    region,
    output_prefix,
    success_csv,
    extra_jsonl_paths,
    book_paths,
    video_info_paths,
    add_video_url_col,
    add_zh_video_url_col,
):
    output_prefix = normalize_prefix(output_prefix)
    keys = list_output_jsonl_keys(bucket, output_prefix, region)
    rows = OrderedDict()

    def add_row(item):
        image_name = item.get("image_name")
        if not image_name:
            return
        contains_child = item.get("contains_child")
        age_group = item.get("age_group")
        is_child_targeted = item.get("is_child_targeted")
        confidence = item.get("confidence")
        if (
            contains_child is None
            or age_group is None
            or is_child_targeted is None
            or confidence is None
        ):
            return
        rows[image_name] = {
            "image_name": image_name,
            "contains_child": contains_child,
            "age_group": age_group,
            "is_child_targeted": is_child_targeted,
            "confidence": confidence,
        }

    s3 = boto3.client("s3", region_name=region)
    for key in keys:
        obj = s3.get_object(Bucket=bucket, Key=key)
        for line in obj["Body"].iter_lines():
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            if record.get("error") or record.get("outputRecord", {}).get("error"):
                continue
            model_output = (
                record.get("modelOutput")
                or record.get("outputRecord", {}).get("modelOutput")
                or {}
            )
            output_body = model_output.get("body")
            if isinstance(output_body, str):
                try:
                    output_body = json.loads(output_body)
                except json.JSONDecodeError:
                    output_body = None
            if isinstance(output_body, dict):
                add_row(output_body)
                continue
            if isinstance(output_body, list):
                for item in output_body:
                    if isinstance(item, dict):
                        add_row(item)
                continue

            text_payload = None
            output_obj = model_output.get("output") or {}
            message = output_obj.get("message") or {}
            content = message.get("content") or []
            if isinstance(content, list) and content:
                first = content[0]
                if isinstance(first, dict):
                    text_payload = first.get("text")
            parsed = parse_text_payload(text_payload)
            if isinstance(parsed, dict):
                add_row(parsed)
            elif isinstance(parsed, list):
                for item in parsed:
                    if isinstance(item, dict):
                        add_row(item)

    for path in extra_jsonl_paths:
        p = Path(path)
        if not p.exists():
            continue
        with p.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    record = json.loads(line)
                except json.JSONDecodeError:
                    continue
                raw_text = record.get("raw_text") or ""
                parsed = parse_text_payload(raw_text)
                if isinstance(parsed, dict):
                    add_row(parsed)
                elif isinstance(parsed, list):
                    for item in parsed:
                        if isinstance(item, dict):
                            add_row(item)

    book_mapping = build_book_mapping(book_paths) if book_paths else {}
    video_info_mapping = build_video_info_mapping(video_info_paths) if video_info_paths else {}
    combined_mapping = {}
    combined_mapping.update(video_info_mapping)
    combined_mapping.update(book_mapping)
    fieldnames = [
        "image_name",
        "contains_child",
        "age_group",
        "is_child_targeted",
        "confidence",
    ]
    if add_zh_video_url_col:
        fieldnames.append("视频链接")
    if add_video_url_col:
        fieldnames.append("video URL")
    if combined_mapping:
        fieldnames.append("UID")
        fieldnames.append("频道名称")

    with open(success_csv, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows.values():
            image_name = row.get("image_name") or ""
            video_id = video_id_from_image(image_name)
            if add_zh_video_url_col:
                row["视频链接"] = (
                    f"https://www.youtube.com/watch?v={video_id}" if video_id else ""
                )
            if add_video_url_col:
                row["video URL"] = (
                    f"https://www.youtube.com/watch?v={video_id}" if video_id else ""
                )
            if combined_mapping:
                channel_id, channel_name = combined_mapping.get(video_id, ("", ""))
                row["UID"] = channel_id
                row["频道名称"] = channel_name
            writer.writerow(row)

    return {"output_files": len(keys), "unique_success": len(rows)}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--bucket", default=os.environ.get("BUCKET", "nova-test-image"))
    parser.add_argument("--input-prefix", default=os.environ.get("INPUT_PREFIX", "input_from_url/"))
    parser.add_argument("--batch-input-prefix", default=os.environ.get("BATCH_INPUT_PREFIX", "batch/input/"))
    parser.add_argument(
        "--batch-image-prefix",
        default=os.environ.get("BATCH_IMAGE_PREFIX", "batch/input/"),
    )
    parser.add_argument("--batch-output-prefix", default=os.environ.get("BATCH_OUTPUT_PREFIX", "output/"))
    parser.add_argument("--use-job-prefix", action="store_true")
    parser.add_argument("--limit", type=int, default=int(os.environ.get("LIMIT", "500")))
    parser.add_argument("--region", default=os.environ.get("AWS_REGION", "ap-northeast-1"))
    parser.add_argument("--test-only", action="store_true")
    parser.add_argument("--skip-copy-existing", action="store_true")
    parser.add_argument("--skip-head-check", action="store_true")
    parser.add_argument("--input-json", action="append", default=[])
    parser.add_argument("--resume-output-prefix", action="append", default=[])
    parser.add_argument("--resume-success-csv", action="append", default=[])
    parser.add_argument("--reuse-flow", action="store_true")
    parser.add_argument(
        "--rebuild-success-csv",
        default=os.environ.get("SUCCESS_CSV", "success_results.csv"),
    )
    parser.add_argument(
        "--rebuild-output-prefix",
        default=os.environ.get("OUTPUT_PREFIX", "output/"),
    )
    parser.add_argument("--extra-success-jsonl", action="append", default=[])
    parser.add_argument("--book-xlsx", action="append", default=[])
    parser.add_argument("--video-info-json", action="append", default=[])
    parser.add_argument("--add-video-url-col", action="store_true")
    parser.add_argument("--add-zh-video-url-col", action="store_true")
    parser.add_argument("--rebuild-only", action="store_true")
    parser.add_argument(
        "--model-id",
        default=os.environ.get(
            "MODEL_ID",
            "apac.amazon.nova-lite-v1:0",
        ),
    )
    parser.add_argument(
        "--role-arn",
        default=os.environ.get(
            "ROLE_ARN",
            "arn:aws:iam::359144475210:role/BedrockBatchInferenceRole",
        ),
    )
    args = parser.parse_args()

    ensure_inference_profile(args.model_id)

    if args.reuse_flow:
        book_paths = [Path(p) for p in args.book_xlsx]
        if not book_paths:
            for name in ("Book1.xlsx", "BOOK2.xlsx"):
                p = Path(name)
                if p.exists():
                    book_paths.append(p)
        stats = rebuild_success_csv(
            args.bucket,
            args.region,
            args.rebuild_output_prefix,
            args.rebuild_success_csv,
            args.extra_success_jsonl,
            book_paths,
            args.video_info_json,
            True if (args.add_video_url_col or args.reuse_flow) else False,
            True if (args.add_zh_video_url_col or args.reuse_flow) else False,
        )
        print(json.dumps(stats, ensure_ascii=False, indent=2))
        if args.rebuild_only:
            return
        if args.rebuild_success_csv not in args.resume_success_csv:
            args.resume_success_csv = [args.rebuild_success_csv] + args.resume_success_csv

    now = dt.datetime.utcnow().strftime("%Y%m%d%H%M%S")
    job_name = f"nova2lite-batch-{now}"

    input_prefix = normalize_prefix(args.input_prefix)
    batch_input_prefix = normalize_prefix(args.batch_input_prefix)
    batch_image_prefix = normalize_prefix(args.batch_image_prefix)
    batch_output_prefix = normalize_prefix(args.batch_output_prefix)
    use_job_prefix = args.use_job_prefix or os.environ.get("USE_JOB_PREFIX", "1") == "1"
    if use_job_prefix:
        batch_job_prefix = normalize_prefix(f"{batch_input_prefix}{job_name}")
        batch_image_prefix = batch_job_prefix
        batch_input_prefix = batch_job_prefix
    if not batch_image_prefix.startswith(batch_input_prefix):
        batch_image_prefix = batch_input_prefix

    if args.input_json:
        keys = load_image_keys_from_json(args.input_json, input_prefix, args.limit)
    else:
        keys = list_image_keys(args.bucket, input_prefix, args.region, args.limit)
    if not keys:
        raise SystemExit("No images found under input prefix.")
    if args.resume_output_prefix:
        resume_prefixes = [normalize_prefix(p) for p in args.resume_output_prefix]
        processed = load_processed_image_names(args.bucket, resume_prefixes, args.region)
        keys = [key for key in keys if Path(key).name not in processed]
        if not keys:
            raise SystemExit("No images remaining after resume filter.")
    if args.resume_success_csv:
        processed = load_processed_from_csv(args.resume_success_csv)
        keys = [key for key in keys if Path(key).name not in processed]
        if not keys:
            raise SystemExit("No images remaining after resume CSV filter.")
    if len(keys) > 50000:
        raise SystemExit(f"Too many records: {len(keys)} exceeds 50000 limit.")

    s3 = boto3.client("s3", region_name=args.region)
    fallback_prefixes = [
        normalize_prefix("input_from_url/"),
        normalize_prefix("batch/input/"),
    ]
    local_path = Path(f"/tmp/batch_input_{job_name}.jsonl")
    with local_path.open("w", encoding="utf-8") as f:
        for idx, key in enumerate(keys):
            image_name = Path(key).name
            image_format = detect_format(key)
            target_key = f"{batch_image_prefix}{image_name}"
            source_key = key
            if not args.skip_head_check:
                candidates = [key]
                for prefix in fallback_prefixes:
                    alt = f"{prefix}{image_name}"
                    if alt not in candidates:
                        candidates.append(alt)
                source_key = None
                for candidate in candidates:
                    try:
                        s3.head_object(Bucket=args.bucket, Key=candidate)
                        source_key = candidate
                        break
                    except Exception:
                        continue
                if not source_key:
                    continue
            if args.skip_copy_existing:
                try:
                    s3.head_object(Bucket=args.bucket, Key=target_key)
                except Exception:
                    try:
                        s3.copy(
                            {"Bucket": args.bucket, "Key": source_key},
                            args.bucket,
                            target_key,
                        )
                    except Exception:
                        continue
            else:
                try:
                    s3.copy(
                        {"Bucket": args.bucket, "Key": source_key},
                        args.bucket,
                        target_key,
                    )
                except Exception:
                    continue
            s3_uri = f"s3://{args.bucket}/{target_key}"
            record = build_record(str(idx), s3_uri, image_name, image_format)
            f.write(json.dumps(record, ensure_ascii=False) + "\n")

    input_key = f"{batch_input_prefix}{local_path.name}"
    s3.upload_file(str(local_path), args.bucket, input_key)

    if args.test_only:
        print(
            json.dumps(
                {
                    "jobName": job_name,
                    "inputS3Uri": f"s3://{args.bucket}/{input_key}",
                    "inputPrefix": f"s3://{args.bucket}/{batch_input_prefix}",
                    "outputS3Uri": f"s3://{args.bucket}/{batch_output_prefix}{job_name}/",
                    "records": len(keys),
                    "modelId": args.model_id,
                    "roleArn": args.role_arn,
                    "testOnly": True,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return

    bedrock = boto3.client("bedrock", region_name=args.region)
    job = bedrock.create_model_invocation_job(
        jobName=job_name,
        roleArn=args.role_arn,
        modelId=args.model_id,
        inputDataConfig={
            "s3InputDataConfig": {
                "s3InputFormat": "JSONL",
                "s3Uri": f"s3://{args.bucket}/{batch_input_prefix}",
            }
        },
        outputDataConfig={"s3OutputDataConfig": {"s3Uri": f"s3://{args.bucket}/{batch_output_prefix}{job_name}/"}},
    )

    print(
        json.dumps(
            {
                "jobName": job_name,
                "jobArn": job.get("jobArn"),
                "inputS3Uri": f"s3://{args.bucket}/{input_key}",
                "inputPrefix": f"s3://{args.bucket}/{batch_input_prefix}",
                "outputS3Uri": f"s3://{args.bucket}/{batch_output_prefix}{job_name}/",
                "records": len(keys),
                "modelId": args.model_id,
                "roleArn": args.role_arn,
                "testOnly": False,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
