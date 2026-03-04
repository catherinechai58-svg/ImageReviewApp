import argparse
import importlib.util
import json
import os
import sys
import urllib.request
from urllib.error import HTTPError, URLError
from pathlib import Path
from xml.etree import ElementTree
import subprocess

import boto3
import openpyxl


def load_items(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def download_image(url):
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        content_type = resp.headers.get("Content-Type", "")
        data = resp.read()
        return data, content_type


def upload_bytes(bucket, key, data, region):
    s3 = boto3.client("s3", region_name=region)
    s3.put_object(Bucket=bucket, Key=key, Body=data)


def load_video_info(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {str(h).strip().lower(): idx for idx, h in enumerate(headers) if h is not None}
    video_idx = header_map.get("video_id")
    channel_idx = header_map.get("channel_id")
    if video_idx is None or channel_idx is None:
        wb.close()
        raise ValueError("video_id or channel_id column not found in video info sheet")
    video_to_channel = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        video_id = row[video_idx]
        channel_id = row[channel_idx]
        if video_id:
            video_to_channel[str(video_id)] = str(channel_id) if channel_id is not None else None
    wb.close()
    return video_to_channel


def load_channel_ids(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {str(h).strip().lower(): idx for idx, h in enumerate(headers) if h is not None}
    uid_idx = header_map.get("uid")
    if uid_idx is None:
        wb.close()
        raise ValueError("UID column not found in channel list sheet")
    channel_ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        uid = row[uid_idx]
        if uid:
            channel_ids.add(str(uid))
    wb.close()
    return channel_ids


def load_channel_map(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {str(h).strip().lower(): idx for idx, h in enumerate(headers) if h is not None}
    uid_idx = header_map.get("uid")
    name_idx = (
        header_map.get("频道名称")
        or header_map.get("channel_name")
        or header_map.get("channel_display_name")
        or header_map.get("name")
    )
    if uid_idx is None:
        wb.close()
        raise ValueError("UID column not found in channel list sheet")
    channel_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        uid = row[uid_idx]
        if not uid:
            continue
        name = row[name_idx] if name_idx is not None else ""
        channel_map[str(uid)] = str(name) if name is not None else ""
    wb.close()
    return channel_map


def parse_channel_ids(text):
    if not text:
        return []
    parts = []
    for chunk in text.replace(",", " ").split():
        val = chunk.strip()
        if val:
            parts.append(val)
    seen = set()
    ordered = []
    for cid in parts:
        if cid in seen:
            continue
        seen.add(cid)
        ordered.append(cid)
    return ordered


def load_hotword_module(path):
    module_path = Path(path)
    hotword_dir = module_path.parent
    scripts_dir = hotword_dir / "scripts"
    if str(hotword_dir) not in sys.path:
        sys.path.insert(0, str(hotword_dir))
    if scripts_dir.exists() and str(scripts_dir) not in sys.path:
        sys.path.insert(0, str(scripts_dir))
    spec = importlib.util.spec_from_file_location("hotword_workflow", str(module_path))
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module, hotword_dir


def get_youtube_service_readonly(hotword_dir, token_file):
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build

    scopes = ["https://www.googleapis.com/auth/youtube.readonly"]
    token_file = Path(token_file)
    client_secret_file = Path(hotword_dir) / "client_secret.json"
    if not client_secret_file.exists():
        raise FileNotFoundError(f"client_secret.json not found in {hotword_dir}")
    creds = None
    if token_file.exists():
        try:
            creds = Credentials.from_authorized_user_file(str(token_file), scopes)
        except Exception:
            creds = None
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception:
                creds = None
        if not creds or not creds.valid:
            flow = InstalledAppFlow.from_client_secrets_file(str(client_secret_file), scopes)
            creds = flow.run_local_server(port=0)
        token_file.parent.mkdir(parents=True, exist_ok=True)
        with token_file.open("w", encoding="utf-8") as f:
            f.write(creds.to_json())
    return build("youtube", "v3", credentials=creds, cache_discovery=False)


def fetch_video_ids_with_hotword(module, hotword_dir, channel_ids, limit_per_channel, token_file):
    cwd = os.getcwd()
    try:
        os.chdir(hotword_dir)
        youtube = None
        if hasattr(module, "get_authenticated_services"):
            try:
                youtube, _ = module.get_authenticated_services()
            except Exception:
                youtube = None
        if youtube is None and hasattr(module, "get_authenticated_service"):
            try:
                youtube = module.get_authenticated_service()
            except Exception:
                youtube = None
        if youtube is None:
            youtube = get_youtube_service_readonly(hotword_dir, token_file)
        video_map = {}
        for channel_id in channel_ids:
            try:
                video_ids = module.list_channel_video_ids(youtube, channel_id)
            except Exception:
                video_ids = []
            if limit_per_channel:
                video_ids = video_ids[:limit_per_channel]
            video_map[channel_id] = video_ids
        return video_map
    finally:
        os.chdir(cwd)


def parse_feed_entries(xml_bytes, channel_id, limit_per_channel, template):
    items = []
    try:
        root = ElementTree.fromstring(xml_bytes)
    except ElementTree.ParseError:
        return items
    for entry in root.iter():
        if not entry.tag.endswith("entry"):
            continue
        video_id = None
        thumb_url = None
        title = None
        channel_name = None
        for child in entry.iter():
            tag = child.tag
            if tag.endswith("videoId"):
                if child.text:
                    video_id = child.text.strip()
            elif tag.endswith("thumbnail"):
                url = child.attrib.get("url")
                if url:
                    thumb_url = url
            elif tag.endswith("title"):
                if child.text:
                    title = child.text.strip()
            elif tag.endswith("name"):
                if child.text:
                    channel_name = child.text.strip()
        if not video_id:
            continue
        cover_url = thumb_url or template.format(video_id=video_id)
        items.append(
            {
                "image_name": f"{video_id}.jpg",
                "url": cover_url,
                "video_id": video_id,
                "channel_id": channel_id,
                "channel_name": channel_name or "",
                "title": title or "",
            }
        )
        if limit_per_channel and len(items) >= limit_per_channel:
            break
    return items


def fetch_video_ids_with_feed(channel_ids, limit_per_channel, timeout, template):
    items = []
    for channel_id in channel_ids:
        feed_url = f"https://www.youtube.com/feeds/videos.xml?channel_id={channel_id}"
        req = urllib.request.Request(feed_url, headers={"User-Agent": "Mozilla/5.0"})
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                data = resp.read()
        except Exception:
            continue
        items.extend(parse_feed_entries(data, channel_id, limit_per_channel, template))
    return items


def build_items_from_channels(
    channel_ids,
    channel_name_map,
    limit_per_channel,
    template,
    hotword_script,
    use_hotword,
    token_file,
):
    if use_hotword and hotword_script:
        module, hotword_dir = load_hotword_module(hotword_script)
        video_map = fetch_video_ids_with_hotword(
            module,
            hotword_dir,
            channel_ids,
            limit_per_channel,
            token_file,
        )
        items = []
        for channel_id, video_ids in video_map.items():
            for video_id in video_ids:
                items.append(
                    {
                        "image_name": f"{video_id}.jpg",
                        "url": template.format(video_id=video_id),
                        "video_id": video_id,
                        "channel_id": channel_id,
                        "channel_name": channel_name_map.get(channel_id, ""),
                    }
                )
        return items
    return fetch_video_ids_with_feed(channel_ids, limit_per_channel, 30, template)


def run_bedrock_workflow(args, output_json, rebuild_only):
    cmd = [
        "python3",
        args.bedrock_script,
        "--reuse-flow",
        "--bucket",
        args.bucket,
        "--region",
        args.region,
        "--input-prefix",
        args.prefix,
        "--input-json",
        output_json,
        "--batch-input-prefix",
        args.batch_input_prefix,
        "--batch-image-prefix",
        args.batch_image_prefix,
        "--batch-output-prefix",
        args.batch_output_prefix,
        "--rebuild-success-csv",
        args.rebuild_success_csv,
        "--rebuild-output-prefix",
        args.rebuild_output_prefix,
        "--video-info-json",
        output_json,
        "--limit",
        str(args.bedrock_limit),
    ]
    if args.use_job_prefix:
        cmd.append("--use-job-prefix")
    if rebuild_only:
        cmd.append("--rebuild-only")
    if args.add_video_url_col:
        cmd.append("--add-video-url-col")
    if args.add_zh_video_url_col:
        cmd.append("--add-zh-video-url-col")
    for path in args.extra_success_jsonl:
        cmd.extend(["--extra-success-jsonl", path])
    for path in args.book_xlsx:
        cmd.extend(["--book-xlsx", path])
    subprocess.run(cmd, check=False)


def filter_items_by_channel(items, video_to_channel, channel_ids):
    filtered = []
    missing_video = []
    channel_mismatch = []
    for item in items:
        image_name = item.get("image_name")
        if not image_name:
            missing_video.append({"image_name": image_name, "reason": "missing_image_name"})
            continue
        video_id = Path(image_name).stem
        channel_id = video_to_channel.get(video_id)
        if channel_id is None:
            missing_video.append({"image_name": image_name, "video_id": video_id, "reason": "video_id_not_found"})
            continue
        if channel_id not in channel_ids:
            channel_mismatch.append({"image_name": image_name, "video_id": video_id, "channel_id": channel_id})
            continue
        filtered.append(item)
    return filtered, missing_video, channel_mismatch


def write_items(path, items):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)


def load_resume_set(path):
    if not path or not os.path.exists(path):
        return set()
    processed = set()
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
                image_name = record.get("image_name")
                if image_name:
                    processed.add(image_name)
            except json.JSONDecodeError:
                continue
    return processed


def append_resume(path, record):
    if not path:
        return
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="/Users/catherine/Documents/trae_projects/aws/lambda_input.json")
    parser.add_argument("--bucket", default=os.environ.get("BUCKET", "nova-test-image"))
    parser.add_argument("--prefix", default=os.environ.get("PREFIX", "input_from_url/"))
    parser.add_argument("--smoke-count", type=int, default=int(os.environ.get("SMOKE_COUNT", "3")))
    parser.add_argument("--region", default=os.environ.get("AWS_REGION", "ap-northeast-1"))
    parser.add_argument("--placeholder-max-bytes", type=int, default=int(os.environ.get("PLACEHOLDER_MAX_BYTES", "2000")))
    parser.add_argument("--resume-file", default=os.environ.get("RESUME_FILE", "upload_resume.jsonl"))
    parser.add_argument("--skip-existing", action="store_true")
    parser.add_argument("--video-info-xlsx", default=os.environ.get("VIDEO_INFO_XLSX", ""))
    parser.add_argument("--channel-list-xlsx", default=os.environ.get("CHANNEL_LIST_XLSX", ""))
    parser.add_argument("--channel-fetch", action="store_true")
    parser.add_argument("--channel-ids", default=os.environ.get("CHANNEL_IDS", ""))
    parser.add_argument("--channel-video-limit", type=int, default=int(os.environ.get("CHANNEL_VIDEO_LIMIT", "0")))
    parser.add_argument(
        "--channel-template",
        default=os.environ.get("CHANNEL_TEMPLATE", "https://i.ytimg.com/vi/{video_id}/mqdefault.jpg"),
    )
    parser.add_argument(
        "--hotword-script",
        default=os.environ.get("HOTWORD_SCRIPT", "/Users/catherine/Downloads/热词云/完整分析-复用版.py"),
    )
    parser.add_argument(
        "--oauth-token-file",
        default=os.environ.get("OAUTH_TOKEN_FILE", "/Users/catherine/Documents/trae_projects/aws/hotword_token.json"),
    )
    parser.add_argument("--use-hotword-script", action="store_true")
    parser.add_argument("--filter-only", action="store_true")
    parser.add_argument("--output-json", default=os.environ.get("OUTPUT_JSON", ""))
    parser.add_argument("--workflow", action="store_true")
    parser.add_argument("--workflow-stage", default=os.environ.get("WORKFLOW_STAGE", "submit"))
    parser.add_argument("--bedrock-script", default=os.environ.get("BEDROCK_SCRIPT", "/Users/catherine/Documents/trae_projects/aws/bedrock_batch_nova2_lite.py"))
    parser.add_argument("--bedrock-limit", type=int, default=int(os.environ.get("BEDROCK_LIMIT", "0")))
    parser.add_argument("--batch-input-prefix", default=os.environ.get("BATCH_INPUT_PREFIX", "batch/input/"))
    parser.add_argument("--batch-image-prefix", default=os.environ.get("BATCH_IMAGE_PREFIX", "batch/input/"))
    parser.add_argument("--batch-output-prefix", default=os.environ.get("BATCH_OUTPUT_PREFIX", "output/"))
    parser.add_argument("--rebuild-success-csv", default=os.environ.get("SUCCESS_CSV", "success_results.csv"))
    parser.add_argument("--rebuild-output-prefix", default=os.environ.get("OUTPUT_PREFIX", "output/"))
    parser.add_argument("--extra-success-jsonl", action="append", default=[])
    parser.add_argument("--book-xlsx", action="append", default=[])
    parser.add_argument("--add-video-url-col", action="store_true")
    parser.add_argument("--add-zh-video-url-col", action="store_true")
    parser.add_argument("--use-job-prefix", action="store_true")
    args = parser.parse_args()

    if args.workflow and args.workflow_stage == "rebuild":
        output_json = args.output_json or args.input
        run_bedrock_workflow(args, output_json, True)
        return

    output_json = args.output_json or args.input
    channel_ids_from_arg = parse_channel_ids(args.channel_ids)
    if args.channel_fetch and (channel_ids_from_arg or args.channel_list_xlsx):
        channel_map = load_channel_map(args.channel_list_xlsx) if args.channel_list_xlsx else {}
        channel_ids = channel_ids_from_arg or list(channel_map.keys())
        if channel_ids_from_arg and channel_map:
            channel_map = {cid: channel_map.get(cid, "") for cid in channel_ids}
        use_hotword = args.use_hotword_script or bool(args.hotword_script)
        items = build_items_from_channels(
            channel_ids,
            channel_map,
            args.channel_video_limit,
            args.channel_template,
            args.hotword_script,
            use_hotword,
            args.oauth_token_file,
        )
        write_items(output_json, items)
    else:
        items = load_items(args.input)
    total = len(items)
    filter_summary = {}
    if args.video_info_xlsx and args.channel_list_xlsx:
        video_to_channel = load_video_info(args.video_info_xlsx)
        channel_ids = load_channel_ids(args.channel_list_xlsx)
        filtered, missing_video, channel_mismatch = filter_items_by_channel(items, video_to_channel, channel_ids)
        filter_summary = {
            "filtered_count": len(filtered),
            "missing_video_count": len(missing_video),
            "channel_mismatch_count": len(channel_mismatch),
        }
        items = filtered

    if args.filter_only:
        if args.output_json:
            write_items(args.output_json, items)
        summary = {
            "input_count": total,
            "filtered_count": len(items),
            "filter_summary": filter_summary,
            "output_json": args.output_json or None,
        }
        json.dump(summary, sys.stdout, ensure_ascii=False, indent=2)
        sys.stdout.write("\n")
        return

    target_items = items[: args.smoke_count] if args.smoke_count > 0 else items

    uploaded = []
    failed = []
    skipped = []
    processed = load_resume_set(args.resume_file)
    s3 = boto3.client("s3", region_name=args.region)

    for item in target_items:
        image_name = item.get("image_name")
        url = item.get("url")
        if not image_name or not url:
            failed.append({"image_name": image_name, "url": url, "error": "missing_fields"})
            append_resume(args.resume_file, {"image_name": image_name, "status": "failed", "error": "missing_fields"})
            continue
        if image_name in processed:
            skipped.append({"image_name": image_name, "reason": "already_processed"})
            continue
        key = f"{args.prefix}{image_name}"
        try:
            if args.skip_existing:
                try:
                    s3.head_object(Bucket=args.bucket, Key=key)
                    skipped.append({"image_name": image_name, "key": key, "reason": "exists"})
                    append_resume(args.resume_file, {"image_name": image_name, "status": "skipped", "reason": "exists"})
                    continue
                except Exception:
                    pass
            data, content_type = download_image(url)
            if content_type and not content_type.startswith("image/"):
                failed.append(
                    {
                        "image_name": image_name,
                        "url": url,
                        "error": "非图片内容",
                        "content_type": content_type,
                    }
                )
                append_resume(
                    args.resume_file,
                    {"image_name": image_name, "status": "failed", "error": "非图片内容", "content_type": content_type},
                )
                continue
            if len(data) <= args.placeholder_max_bytes:
                failed.append(
                    {
                        "image_name": image_name,
                        "url": url,
                        "error": "灰色占位图/不可用缩略图",
                        "size_bytes": len(data),
                        "content_type": content_type,
                    }
                )
                append_resume(
                    args.resume_file,
                    {"image_name": image_name, "status": "failed", "error": "灰色占位图/不可用缩略图"},
                )
                continue
            s3.put_object(Bucket=args.bucket, Key=key, Body=data)
            uploaded.append({"image_name": image_name, "key": key})
            append_resume(args.resume_file, {"image_name": image_name, "status": "uploaded", "key": key})
        except (HTTPError, URLError, TimeoutError, ValueError) as e:
            failed.append({"image_name": image_name, "url": url, "error": str(e)})
            append_resume(args.resume_file, {"image_name": image_name, "status": "failed", "error": str(e)})
        except Exception as e:
            failed.append({"image_name": image_name, "url": url, "error": str(e)})
            append_resume(args.resume_file, {"image_name": image_name, "status": "failed", "error": str(e)})

    summary = {
        "bucket": args.bucket,
        "prefix": args.prefix,
        "region": args.region,
        "input_count": total,
        "filter_summary": filter_summary,
        "attempted": len(target_items),
        "uploaded": uploaded,
        "failed": failed,
        "skipped": skipped,
        "placeholder_max_bytes": args.placeholder_max_bytes,
        "resume_file": args.resume_file,
        "skip_existing": args.skip_existing,
    }
    json.dump(summary, sys.stdout, ensure_ascii=False, indent=2)
    sys.stdout.write("\n")
    if args.workflow and args.workflow_stage == "submit" and not args.filter_only:
        run_bedrock_workflow(args, output_json, False)


if __name__ == "__main__":
    main()
